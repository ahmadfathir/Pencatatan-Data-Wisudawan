from flask import Flask, render_template, request, redirect, url_for, send_file
import csv
import os
import traceback
from datetime import datetime

# Eksplisit path ke folder templates dan static
BASE_DIR = os.path.dirname(os.path.abspath(__file__))
app = Flask(
    __name__,
    template_folder=os.path.join(BASE_DIR, "templates"),
    static_folder=os.path.join(BASE_DIR, "static")
)

# Vercel: hanya /tmp yang writable
DATA_DIR = "/tmp/wisuda_data"
CSV_FILE = os.path.join(DATA_DIR, "wisudawan.csv")

FIELDNAMES = [
    "id", "tanggal_daftar", "nama_lengkap", "nim",
    "fakultas", "program_studi", "nomor_hp", "email", "catatan"
]


def ensure_csv():
    os.makedirs(DATA_DIR, exist_ok=True)
    if not os.path.exists(CSV_FILE):
        with open(CSV_FILE, mode="w", newline="", encoding="utf-8") as f:
            csv.writer(f).writerow(FIELDNAMES)


def read_data():
    ensure_csv()
    with open(CSV_FILE, mode="r", encoding="utf-8") as f:
        return list(csv.DictReader(f))


def write_data(data):
    ensure_csv()
    with open(CSV_FILE, mode="w", newline="", encoding="utf-8") as f:
        w = csv.DictWriter(f, fieldnames=FIELDNAMES)
        w.writeheader()
        w.writerows(data)


# ── Error handler: tampilkan detail error (bantu debug) ──
@app.errorhandler(Exception)
def handle_exception(e):
    tb = traceback.format_exc()
    return f"<pre style='padding:2rem;font-size:13px'><b>ERROR:</b>\n{tb}</pre>", 500


@app.route("/")
def index():
    data = read_data()
    data = sorted(data, key=lambda x: int(x["id"]), reverse=True)
    return render_template("index.html", data=data)


@app.route("/tambah", methods=["GET", "POST"])
def tambah():
    if request.method == "POST":
        data = read_data()
        next_id = max((int(r["id"]) for r in data), default=0) + 1
        data.append({
            "id": str(next_id),
            "tanggal_daftar": datetime.now().strftime("%Y-%m-%d %H:%M"),
            "nama_lengkap": request.form.get("nama_lengkap", "").strip(),
            "nim":          request.form.get("nim", "").strip(),
            "fakultas":     request.form.get("fakultas", "").strip(),
            "program_studi":request.form.get("program_studi", "").strip(),
            "nomor_hp":     request.form.get("nomor_hp", "").strip(),
            "email":        request.form.get("email", "").strip(),
            "catatan":      request.form.get("catatan", "").strip(),
        })
        write_data(data)
        return redirect(url_for("index"))
    return render_template("tambah.html")


@app.route("/hapus/<int:id>")
def hapus(id):
    data = [r for r in read_data() if int(r["id"]) != id]
    write_data(data)
    return redirect(url_for("index"))


@app.route("/export")
def export_csv():
    data = read_data()
    if not data:
        return redirect(url_for("index"))

    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter

    wb = Workbook()
    ws = wb.active
    ws.title = "Wisudawan"

    headers = ["ID", "Tanggal Daftar", "Nama Lengkap", "NIM",
               "Fakultas", "Program Studi", "Nomor HP", "Email", "Catatan"]

    thin    = Side(style="thin", color="DDDDDD")
    border  = Border(left=thin, right=thin, top=thin, bottom=thin)

    ws.append(headers)
    for cell in ws[1]:
        cell.fill = PatternFill("solid", fgColor="1A1208")
        cell.font = Font(bold=True, color="D9B96A")
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = border

    for row in data:
        ws.append([row["id"], row["tanggal_daftar"], row["nama_lengkap"],
                   row["nim"], row["fakultas"], row["program_studi"],
                   row["nomor_hp"], row["email"], row["catatan"]])

    for i, col in enumerate(ws.columns, 1):
        w = max((len(str(c.value or "")) for c in col), default=8)
        ws.column_dimensions[get_column_letter(i)].width = min(w + 3, 40)
        for cell in col[1:]:
            cell.border = border
            cell.alignment = Alignment(vertical="top", wrap_text=True)

    ws.freeze_panes = "A2"
    ws.auto_filter.ref = ws.dimensions

    out = os.path.join(DATA_DIR, "data_wisudawan.xlsx")
    wb.save(out)

    return send_file(out, as_attachment=True,
                     download_name="data_wisudawan.xlsx",
                     mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")


if __name__ == "__main__":
    app.run(debug=True)
